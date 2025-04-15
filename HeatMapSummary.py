# Final version: City/Portal Cores with alignment + auto-fit column widths + full borders (single Story column)

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
import os

# === Define Tower and File Names ===
tower_groups = [
    {
        "name": "City Cores",
        "towers": ["N1", "N2", "N3", "N4", "S1", "S2", "S3", "S4"],
        "col_span": 30,
        "layout": "4-over-4"
    },
    {
        "name": "Portal Cores",
        "towers": ["P1", "P2", "P3", "P4"],
        "col_span": 39,
        "layout": "2-over-2"
    }
]

base_filename = "20250411_M46_JV3.1_{}.xlsm"
source_sheet = "HeatMap"
target_file = "Summary.xlsx"

# === Define Data Number and Data Ranges ===
target_start_row = 3
target_start_col = 2
data_row_count = 113

fields = [
    {
        "name": "V_DCR",
        "sheet": "SummaryVDCR",
        "range_by_group": {
            "City Cores": ("CS3", "DU114"),
            "Portal Cores": ("DR3", "FC115")
        }
    },
    {
        "name": "Vmax_DCR",
        "sheet": "SummaryVmax",
        "range_by_group": {
            "City Cores": ("DX3", "EZ114"),
            "Portal Cores": ("FE3", "GP115")
        }
    },
    {
        "name": "PC_DCR",
        "sheet": "SummaryPCDCR",
        "range_by_group": {
            "City Cores": ("FC3", "GE114"),
            "Portal Cores": ("GR3", "IC115")
        }
    },
    {
        "name": "PT_DCR",
        "sheet": "SummaryPTDCR",
        "range_by_group": {
            "City Cores": ("GH3", "HJ114"),
            "Portal Cores": ("IE3", "JP115")
        }
    },
]

# === Format Settings ===
thin_border = Border(
    left=Side(style='thin', color='AAAAAA'),
    right=Side(style='thin', color='AAAAAA'),
    top=Side(style='thin', color='AAAAAA'),
    bottom=Side(style='thin', color='AAAAAA')
)
center_align = Alignment(horizontal='center', vertical='center')

# === Load or Create Workbook ===
if os.path.exists(target_file):
    print("📄 Loading existing Summary.xlsx...")
    tgt_wb = load_workbook(target_file)
else:
    print("🆕 Creating new Summary.xlsx...")
    tgt_wb = Workbook()
    default_sheet = tgt_wb.active
    tgt_wb.remove(default_sheet)


for field in fields:
    sheet_name = field["sheet"]
    if sheet_name in tgt_wb.sheetnames:
        tgt_ws = tgt_wb[sheet_name]

        # ✅ 清空 A~GV 區塊的內容與格式，保留 GW 欄之後
        for row in tgt_ws.iter_rows(min_row=1, max_row=tgt_ws.max_row, min_col=1, max_col=204):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                cell.border = None
                cell.alignment = None
                cell.font = None
                cell.fill = None

    else:
        tgt_ws = tgt_wb.create_sheet(title=sheet_name)

    current_col = target_start_col


    for group in tower_groups:
        group_name = group["name"]
        towers = group["towers"]
        col_span = group["col_span"]
        layout = group["layout"]
        towers_per_row = 4 if layout == "4-over-4" else 2
        source_range = field["range_by_group"][group_name]
        
        # === Create Total Columns for each Group ===
        group_width = towers_per_row * (col_span + 1) - 1
        tgt_ws.merge_cells(start_row=1, start_column=current_col,
                           end_row=1, end_column=current_col + group_width - 1)
        tgt_ws.cell(row=1, column=current_col).value = group_name

        for row_idx in range(2):
            # === Paste Data of each Tower ===
            row_base = target_start_row + row_idx * (data_row_count + 2)
            row_towers = towers[row_idx * towers_per_row:(row_idx + 1) * towers_per_row]

            for i, tower in enumerate(row_towers):
                col_base = current_col + i * (col_span + 1)
                file_path = base_filename.format(tower)

                if not os.path.exists(file_path):
                    print(f"🔸 Skipping {tower} (file not found)")
                    continue

                try:
                    # === Open Source File and Read Data ===
                    src_wb = load_workbook(file_path, data_only=True, read_only=True)
                    src_ws = src_wb[source_sheet]
                    dcr_values = src_ws[source_range[0]:source_range[1]]
                except Exception as e:
                    print(f"❌ Error reading {tower}: {e}")
                    continue
                
                # === Paste Tower Title ===
                tgt_ws.cell(row=row_base, column=col_base + 1).value = f"Tower {tower}"

                # === Write DCRs and Formats ===
                for i_row, row in enumerate(dcr_values):
                    for j, cell in enumerate(row):
                        val = cell.value
                        if isinstance(val, (float, int)):
                            val = round(val, 2)
                        tgt_ws.cell(row=row_base + 1 + i_row, column=col_base + 1 + j).value = val

                for r in range(row_base + 1, row_base + 1 + data_row_count):
                    for c in range(col_base, col_base + col_span):
                        cell = tgt_ws.cell(row=r, column=c)
                        cell.alignment = center_align
                        cell.border = thin_border

                
                # === Formatting Color Scale for DCRs ===
                pier_start = col_base + 1
                pier_end = col_base + col_span -1
                tgt_ws.conditional_formatting.add(
                    f"{tgt_ws.cell(row=row_base + 1, column=pier_start).coordinate}:{tgt_ws.cell(row=row_base + data_row_count, column=pier_end).coordinate}",
                    ColorScaleRule(
                        start_type='num', start_value=0.6, start_color='C6E0B4',
                        mid_type='num', mid_value=0.8, mid_color='FFEB84',
                        end_type='num', end_value=1.05, end_color='F8696B'
                    )
                )

                print(f"✅ Finished: {field['name']} - {tower}")

        current_col += group_width + 1

    # === Autofit Column Width ===
    for col_cells in tgt_ws.columns:
        first_cell = next((cell for cell in col_cells if not isinstance(cell, MergedCell)), None)
        if first_cell is None:
            continue
        col_letter = get_column_letter(first_cell.column)
        max_len = 0
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        tgt_ws.column_dimensions[col_letter].width = max_len + 1

# === Save workbook ===
tgt_wb.save(target_file)
print("🎉 Summary.xlsx Completed！")
